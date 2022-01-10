import pandas as pd

OMRCOMPRAS_OCD = pd.read_pickle(r'..\OMR_COMPRAS_OCD.pkl')
OMRCOMPRAS_POP = pd.read_pickle(r'..\OMR_COMPRAS_POP.pkl')
NUMANOMESOCD = int(open('../NUMANOMESOCD.txt','r').read())

dimensoes = ['DESDRTCLLATU', 'DESCTGPRD', 'DESDIVFRN', 'NOMGRPECOFRN', 'DESCLLCMPATU']
valores = ['VLRVNDFATLIQ', 'VLRRCTLIQAPU', 'VLRMRGBRT', 'VLRMRGCRB']

DIRETORIA_OCD = OMRCOMPRAS_OCD.groupby(dimensoes)[valores].sum().div(1000).reset_index()
DIRETORIA_OCD.columns = ['DESDRTCLLATU', 'DESCTGPRD', 'DESDIVFRN', 'NOMGRPECOFRN', 'DESCLLCMPATU', 'FAT_OCD', 'RL_OCD', 'MB_OCD', 'MC_OCD']

DIRETORIA_POP = OMRCOMPRAS_POP.groupby(dimensoes)[valores].sum().div(1000).reset_index()
DIRETORIA_POP.columns = ['DESDRTCLLATU', 'DESCTGPRD', 'DESDIVFRN', 'NOMGRPECOFRN', 'DESCLLCMPATU', 'FAT_POP', 'RL_POP', 'MB_POP', 'MC_POP']

DIRETORIA = pd.merge(DIRETORIA_OCD, DIRETORIA_POP, how='left', on=['DESDRTCLLATU', 'DESCTGPRD', 'DESDIVFRN', 'NOMGRPECOFRN', 'DESCLLCMPATU'])
DIRETORIA = DIRETORIA.fillna(0)
DIRETORIA['PERMB'] = DIRETORIA['MB_POP'].div(DIRETORIA['RL_POP'])
DIRETORIA.loc[DIRETORIA.RL_POP == 0,'PERMB'] = 0.23
DIRETORIA = DIRETORIA.sort_values(by='FAT_POP', ascending=False)

ESTADO_OCD = OMRCOMPRAS_OCD.groupby(['CODESTUNI'])[valores].sum().div(1000).reset_index()
ESTADO_OCD.columns = ['CODESTUNI', 'FAT_OCD', 'RL_OCD', 'MB_OCD', 'MC_OCD']

ESTADO_POP = OMRCOMPRAS_POP.groupby(['CODESTUNI'])[valores].sum().div(1000).reset_index()
ESTADO_POP.columns = ['CODESTUNI', 'FAT_POP', 'RL_POP', 'MB_POP', 'MC_POP']

ESTADO = pd.merge(ESTADO_OCD, ESTADO_POP, how='inner', on=['CODESTUNI'])
ESTADO['PERMB'] = ESTADO['MB_POP'].div(ESTADO['RL_POP'])
ESTADO.loc[ESTADO.RL_POP == 0,'PERMB'] = 0.23
ESTADO = ESTADO.sort_values(by='FAT_POP', ascending=False)

#Atualiza arquivo excel
print('iniciando atualização do arquivo xlsx...')
from openpyxl import load_workbook
book = load_workbook(r'..\Plan_Operacional.xlsx')
writer = pd.ExcelWriter(r'..\Plan_Operacional.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['RESUMO DIRETORIA']
ws['A1'].value = NUMANOMESOCD

planilha = 'VAREJO'
ws = book[planilha]
for row in ws['B8:I1000']:
  for cell in row:
    cell.value = None
for row in ws['M8:M1000']:
  for cell in row:
    cell.value = None
for row in ws['S8:S1000']:
  for cell in row:
    cell.value = None
dfxls = DIRETORIA.query('DESDRTCLLATU=="VAREJO ALIMENTAR E FARMA"')[['DESCTGPRD','DESDIVFRN','NOMGRPECOFRN','DESCLLCMPATU', 'FAT_OCD', 'RL_OCD', 'MB_OCD', 'MC_OCD']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=1, startrow=7, header=False, index=False)
dfxls = DIRETORIA.query('DESDRTCLLATU=="VAREJO ALIMENTAR E FARMA"')[['FAT_POP']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=12, startrow=7, header=False, index=False)
dfxls = DIRETORIA.query('DESDRTCLLATU=="VAREJO ALIMENTAR E FARMA"')[['PERMB']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=18, startrow=7, header=False, index=False)

planilha = 'ELETRO'
ws = book[planilha]
for row in ws['B8:I1000']:
  for cell in row:
    cell.value = None
for row in ws['M8:M1000']:
  for cell in row:
    cell.value = None
for row in ws['S8:S1000']:
  for cell in row:
    cell.value = None
dfxls = DIRETORIA.query('DESDRTCLLATU=="ELETRO"')[['DESCTGPRD','DESDIVFRN','NOMGRPECOFRN','DESCLLCMPATU', 'FAT_OCD', 'RL_OCD', 'MB_OCD', 'MC_OCD']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=1, startrow=7, header=False, index=False)
dfxls = DIRETORIA.query('DESDRTCLLATU=="ELETRO"')[['FAT_POP']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=12, startrow=7, header=False, index=False)
dfxls = DIRETORIA.query('DESDRTCLLATU=="ELETRO"')[['PERMB']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=18, startrow=7, header=False, index=False)

planilha = 'MARTCON'
ws = book[planilha]
for row in ws['B8:I1000']:
  for cell in row:
    cell.value = None
for row in ws['M8:M1000']:
  for cell in row:
    cell.value = None
for row in ws['S8:S1000']:
  for cell in row:
    cell.value = None
dfxls = DIRETORIA.query('DESDRTCLLATU=="MARTCON/AGROVET"')[['DESCTGPRD','DESDIVFRN','NOMGRPECOFRN','DESCLLCMPATU', 'FAT_OCD', 'RL_OCD', 'MB_OCD', 'MC_OCD']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=1, startrow=7, header=False, index=False)
dfxls = DIRETORIA.query('DESDRTCLLATU=="MARTCON/AGROVET"')[['FAT_POP']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=12, startrow=7, header=False, index=False)
dfxls = DIRETORIA.query('DESDRTCLLATU=="MARTCON/AGROVET"')[['PERMB']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=18, startrow=7, header=False, index=False)

planilha = 'UF'
ws = book[planilha]
for row in ws['B8:F34']:
  for cell in row:
    cell.value = None
for row in ws['J8:J34']:
  for cell in row:
    cell.value = None
for row in ws['P8:P34']:
  for cell in row:
    cell.value = None
dfxls = ESTADO[['CODESTUNI', 'FAT_OCD', 'RL_OCD', 'MB_OCD', 'MC_OCD']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=1, startrow=7, header=False, index=False)
dfxls = ESTADO['FAT_POP']
dfxls.to_excel(writer, sheet_name=planilha, startcol=9, startrow=7, header=False, index=False)
dfxls = ESTADO['PERMB']
dfxls.to_excel(writer, sheet_name=planilha, startcol=15, startrow=7, header=False, index=False)

writer.save()
book.close()

print("Plan_Operacional.xlsx atualizado!")