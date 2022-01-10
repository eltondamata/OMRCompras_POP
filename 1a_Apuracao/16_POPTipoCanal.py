#Atualizar valor de referencia para calculo do POP ajustado por Tipo de Canal (POP_TIPCNLVND.xlsx)
#Valor de referencia refere-se ao valor definido por fornecedor x UF (Recebido de vendas) para as demais dimensoes do OMR Compras.
import pandas as pd

with open('../Parametros/caminho.txt','r') as f:
    caminho = f.read()

pd.options.display.float_format = '{:,.2f}'.format
#importa base completa OMR_COMPRAS
df = pd.read_feather(caminho + 'bd/OMR_COMPRAS_POP.ft')

#Agrupa valores da Planilha em Diretoria x Categoria
valores = 'VLRVNDFATLIQ VLRRCTLIQAPU VLRMRGBRT VLRMRGCRB VLRCSTMC'.split()
dfcnl = df.groupby('DESTIPCNLVNDOMR')[valores].sum().reset_index()
print('POP TIPO CANAL')
print(dfcnl.to_string())

#Atualiza arquivo excel
print('iniciando atualização do arquivo xlsx...')
from openpyxl import load_workbook
book = load_workbook(caminho + 'POP_TIPCNLVND.xlsx')
writer = pd.ExcelWriter(caminho + 'POP_TIPCNLVND.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

planilha = 'Ajuste'
ws = book[planilha]
for row in ws['A2:F4']:
  for cell in row:
    cell.value = None
dfcnl.to_excel(writer, sheet_name=planilha, startcol=0, startrow=1, header=True, index=False)

writer.save()
book.close()

print("POP_TIPCNLVND.xlsx atualizado!")