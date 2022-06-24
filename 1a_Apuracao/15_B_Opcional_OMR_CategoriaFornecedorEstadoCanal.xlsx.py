#elton.mata@martins.com.br
'''
ATENÇÃO!!! Não é necessário rodar esse processo caso não tenha necessidade de definir a meta no detalhe do cruzmento ( Categoria x Fornecedor x Estado x Canal)
O processo para retorno das metas definidas nesse arquivo é: "22_B_Opcional_BreakBackOMR_COMPRAS_Final.py"

Atualiza arquivo OMR_CategoriaFornecedorEstadoCanal.xlsx com meta POP
Arquivo usado para ajustar a meta nos cruzamento Categoria x Fornecedor x Estado x Canal (Opcional)
Caso não seja necessário ajustes nesse nível de detalhe não é necessário executar esse processo
'''

import pandas as pd

with open('../Parametros/caminho.txt','r') as f:
    caminho = f.read()
    
omrcmp = pd.read_feather(caminho + 'bd/OMR_COMPRAS_POP.ft')
omrcmp.rename(columns={'VLRVNDFATLIQ':'FAT_OCD', 'VLRRCTLIQAPU':'RL_OCD', 'VLRMRGBRT':'MB_OCD', 'VLRCSTMC':'CSTMC_OCD', 'VLRMRGCRB':'MC_OCD'}, inplace=True)

dimensoes = ['DESDRTCLLATU', 'DESCTGPRD', 'CODGRPPRD', 'CODCTGPRD', 'DESDIVFRN', 'CODDIVFRN', 'NOMGRPECOFRN', 'DESCLLCMPATU', 'CODESTUNI', 'DESTIPCNLVNDOMR']
valores = ['FAT_OCD', 'RL_OCD', 'MB_OCD', 'CSTMC_OCD', 'MC_OCD']

#Agrupa dados por diretoria, categoria, forn, grupo forn, celula
df = omrcmp.groupby(dimensoes)[valores].sum().div(1000).reset_index()
df['PERMB'] = df['MB_OCD'].div(df['RL_OCD'])
df.eval('PERMB=MB_OCD/RL_OCD', inplace=True)
df.eval('FATRL=FAT_OCD/RL_OCD', inplace=True)
df.eval('PERCSTMC=CSTMC_OCD/RL_OCD', inplace=True)
df.loc[df.RL_OCD == 0,'PERMB'] = 0.23
df.loc[df.PERMB > 0.6,'PERMB'] = 0.6
df = df.sort_values(by='FAT_OCD', ascending=False)
df.loc[df.DESTIPCNLVNDOMR == "OUTROS CANAIS", 'DESTIPCNLVNDOMR'] = "ATACADO"

#Atualiza arquivo excel
print('iniciando atualização do arquivo xlsx...')
from openpyxl import load_workbook
book = load_workbook(caminho + 'OMR_CategoriaFornecedorEstadoCanal.xlsx')
writer = pd.ExcelWriter(caminho + 'OMR_CategoriaFornecedorEstadoCanal.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
planilha = 'VAREJO'
ws = book[planilha]
for row in ws['B8:N20000']:
  for cell in row:
    cell.value = None
dfxls = df.query('DESDRTCLLATU=="VAREJO ALIMENTAR E FARMA"')[['DESCTGPRD', 'CODGRPPRD', 'CODCTGPRD','DESDIVFRN', 'CODDIVFRN', 'NOMGRPECOFRN','DESCLLCMPATU', 'CODESTUNI', 'DESTIPCNLVNDOMR',  'FAT_OCD', 'PERMB', 'FATRL', 'PERCSTMC']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=1, startrow=7, header=False, index=False)
planilha = 'ELETRO'
ws = book[planilha]
for row in ws['B8:N20000']:
  for cell in row:
    cell.value = None
dfxls = df.query('DESDRTCLLATU=="ELETRO"')[['DESCTGPRD', 'CODGRPPRD', 'CODCTGPRD','DESDIVFRN', 'CODDIVFRN', 'NOMGRPECOFRN','DESCLLCMPATU', 'CODESTUNI', 'DESTIPCNLVNDOMR',  'FAT_OCD', 'PERMB', 'FATRL', 'PERCSTMC']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=1, startrow=7, header=False, index=False)
planilha = 'MARTCON'
ws = book[planilha]
for row in ws['B8:N20000']:
  for cell in row:
    cell.value = None
dfxls = df.query('DESDRTCLLATU=="MARTCON/AGROVET"')[['DESCTGPRD', 'CODGRPPRD', 'CODCTGPRD','DESDIVFRN', 'CODDIVFRN', 'NOMGRPECOFRN','DESCLLCMPATU', 'CODESTUNI', 'DESTIPCNLVNDOMR',  'FAT_OCD', 'PERMB', 'FATRL', 'PERCSTMC']]
dfxls.to_excel(writer, sheet_name=planilha, startcol=1, startrow=7, header=False, index=False)
writer.save()
book.close()

print("OMR_CategoriaFornecedorEstadoCanal.xlsx atualizado!")
