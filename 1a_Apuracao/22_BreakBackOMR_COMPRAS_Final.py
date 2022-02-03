#Gera carga OMR Compras POP com alinhamento das contribuições de CATEGORIAxFORN com ESTADO
#Tempo processamento 52 segundos (BreakBack de 20x, Tambanho da Base Completa = 66.145 linhas)
#log mostrado no terminal de execução e exportado para excel

#Importa planilhas de carga Diretoria x Categoria x Fornecedor
import pandas as pd
import numpy as np
pd.options.display.float_format = '{:,.2f}'.format

with open('../Parametros/caminho.txt','r') as f:
    caminho = f.read()

df = pd.read_excel(caminho + 'Plan_Operacional_CARGA.xlsx', 'VAREJO', skiprows=6, usecols="B:C,M:N,Q:R")
df.insert(0,'DESDRTCLLATU','VAREJO ALIMENTAR E FARMA')
df1 = pd.read_excel(caminho + 'Plan_Operacional_CARGA.xlsx', 'ELETRO', skiprows=6, usecols="B:C,M:N,Q:R")
df1.insert(0,'DESDRTCLLATU','ELETRO')
df = df.append(df1)
df1 = pd.read_excel(caminho + 'Plan_Operacional_CARGA.xlsx', 'MARTCON', skiprows=6, usecols="B:C,M:N,Q:R")
df1.insert(0,'DESDRTCLLATU','MARTCON/AGROVET')
df = df.append(df1)
df.columns = ['DESDRTCLLATU', 'DESCTGPRD', 'DESDIVFRN', 'VLRVNDFATLIQ', 'VLRRCTLIQAPU', 'VLRMRGBRT', 'VLRMRGCRB']
df[['VLRVNDFATLIQ','VLRRCTLIQAPU','VLRMRGBRT', 'VLRMRGCRB']] = df[['VLRVNDFATLIQ','VLRRCTLIQAPU','VLRMRGBRT', 'VLRMRGCRB']].mul(1000)
df['VLRCSTMC'] = df['VLRMRGCRB'] - df['VLRMRGBRT']

df = pd.melt(
	df, id_vars=['DESDRTCLLATU', 'DESCTGPRD', 'DESDIVFRN'],
	value_vars=['VLRVNDFATLIQ','VLRRCTLIQAPU','VLRMRGBRT', 'VLRCSTMC'],
	var_name='MEDIDA',
	value_name='VALOR')

#importa base completa OMR_COMPRAS
df_full = pd.read_feather(caminho + 'bd/OMR_COMPRAS_POPAJT.ft')
df_full = pd.melt(
	df_full, id_vars=['NOMMES', 'CODGRPPRD', 'CODCTGPRD', 'CODSUBCTGPRD', 'CODDIVFRN', 'DESDIVFRN', 'CODESTUNI', 'DESTIPCNLVNDOMR', 'DESCTGPRD', 'DESDRTCLLATU'],
	value_vars=['VLRVNDFATLIQ','VLRRCTLIQAPU','VLRMRGBRT', 'VLRCSTMC'],
	var_name='MEDIDA',
	value_name='DRIVER')
print('\n', "BASE COMPLETA ANTES DO BREAKBACK")
print(df_full.pivot_table(index=['DESDRTCLLATU','MEDIDA'], values='DRIVER', aggfunc=sum).unstack().to_string())

#Base Completa após Distribuir o valor de DIRETORIA x CATEGORIA x FORN na base completa
df_full['DRIVER'] = (df_full['DRIVER'] / df_full.groupby(['DESDRTCLLATU','DESCTGPRD','DESDIVFRN','MEDIDA'])['DRIVER'].transform('sum'))
df_full = df_full.merge(df, how='inner', on=['DESDRTCLLATU','DESCTGPRD','DESDIVFRN','MEDIDA'])
df_full['DRIVER'] = df_full['DRIVER'] * df_full['VALOR']
del df_full['VALOR']
print('\n', "BASE COMPLETA APOS DISTRIBUIR DIRETORIA x CATEGORIA X FORN")
print(df_full.pivot_table(index=['DESDRTCLLATU','MEDIDA'], values='DRIVER', aggfunc=sum).unstack().to_string())

#Agrupa valores da Planilha em Diretoria x Categoria
df_drtctg = df.groupby(['DESDRTCLLATU','DESCTGPRD','MEDIDA'])['VALOR'].sum()
df_drtctg = df_drtctg.reset_index()
#Total por Fornecedor
df_frn = df_full.groupby(['CODDIVFRN','MEDIDA'])['DRIVER'].sum()
df_frn = df_frn.reset_index()
df_frn.columns = ['CODDIVFRN','MEDIDA','VALOR']

#importa planilha por UF
df_uf = pd.read_excel(caminho + 'Plan_Operacional_CARGA.xlsx', 'UF', skiprows=6, usecols="B,J:K,N:O").dropna()
df_uf.columns = ['CODESTUNI', 'VLRVNDFATLIQ', 'VLRRCTLIQAPU', 'VLRMRGBRT', 'VLRMRGCRB']
df_uf['VLRCSTMC'] = df_uf['VLRMRGCRB'] - df_uf['VLRMRGBRT'] 
df_uf = pd.melt(
	df_uf, id_vars=['CODESTUNI'],
	value_vars=['VLRVNDFATLIQ','VLRRCTLIQAPU','VLRMRGBRT', 'VLRCSTMC'],
	var_name='MEDIDA',
	value_name='VALOR')
df_uf['VALOR'] = df_uf['VALOR'].mul(1000)
print('\n', "PLANILHA DE CONTRIBUICAO POR ESTADO")
print(df_uf.pivot_table(index=['MEDIDA'], values='VALOR', aggfunc=sum).unstack().unstack().to_string())

#importa planilha por Tipo Canal
df_tipcnl = pd.read_excel(caminho + 'POP_TIPCNLVND.xlsx', 'POP_TIPCNLVND', usecols = "A:F")
df_tipcnl = pd.melt(
	df_tipcnl, id_vars=['DESTIPCNLVNDOMR'],
	value_vars=['VLRVNDFATLIQ','VLRRCTLIQAPU','VLRMRGBRT', 'VLRCSTMC'],
	var_name='MEDIDA',
	value_name='VALOR')
#df_tipcnl['VALOR'] = df_tipcnl['VALOR'].mul(1000)
print('\n', "POP_TIPCNLVND.xlsx")
print(df_tipcnl.pivot_table(index=['MEDIDA'], values='VALOR', aggfunc=sum).unstack().unstack().to_string())

#INICIO BREAKBACK
for x in range(50):
	#Distribui o valor de ESTADO
	df_full['DRIVER'] = (np.float64(df_full['DRIVER']) / df_full.groupby(['CODESTUNI','MEDIDA'])['DRIVER'].transform('sum'))
	df_full = df_full.merge(df_uf, how='inner', on=['CODESTUNI','MEDIDA'])
	df_full['DRIVER'] = df_full['DRIVER'] * df_full['VALOR']
	del df_full['VALOR']

	#Distribui Tipo Canal
	df_full['DRIVER'] = (np.float64(df_full['DRIVER']) / df_full.groupby(['DESTIPCNLVNDOMR','MEDIDA'])['DRIVER'].transform('sum'))
	df_full = df_full.merge(df_tipcnl, how='inner', on=['DESTIPCNLVNDOMR','MEDIDA'])
	df_full['DRIVER'] = df_full['DRIVER'] * df_full['VALOR']
	del df_full['VALOR']

	#Distribui FORNECEDOR
	df_full['DRIVER'] = (np.float64(df_full['DRIVER']) / df_full.groupby(['CODDIVFRN','MEDIDA'])['DRIVER'].transform('sum'))
	df_full = df_full.merge(df_frn, how='inner', on=['CODDIVFRN','MEDIDA'])
	df_full['DRIVER'] = df_full['DRIVER'] * df_full['VALOR']
	del df_full['VALOR']

	#Distribui DIRETORIA x CATEGORIA
	df_full['DRIVER'] = (np.float64(df_full['DRIVER']) / df_full.groupby(['DESDRTCLLATU','DESCTGPRD','MEDIDA'])['DRIVER'].transform('sum'))
	df_full = df_full.merge(df_drtctg, how='inner', on=['DESDRTCLLATU','DESCTGPRD','MEDIDA'])
	df_full['DRIVER'] = df_full['DRIVER'] * df_full['VALOR']
	del df_full['VALOR']

print('\n', "==> PLANILHA CONTRIBUICAO DIRETORIA X CATEGORIA X FORN")
confere = df.pivot_table(index=['DESDRTCLLATU','MEDIDA'], values='VALOR', aggfunc=sum).unstack()
confere.columns = confere.columns.get_level_values(1)
confere.loc[:,'MC'] = confere['VLRMRGBRT'] + confere['VLRCSTMC']
confere.loc['TOTAL',:] = confere.sum()
print(confere[['VLRVNDFATLIQ', 'VLRRCTLIQAPU', 'VLRMRGBRT', 'VLRCSTMC', 'MC']].to_string())

print('\n', "==> BASE CARGA POR DIRETORIA")
confere2 = df_full.pivot_table(index=['DESDRTCLLATU','MEDIDA'], values='DRIVER', aggfunc=sum).unstack()
confere2.columns = confere2.columns.get_level_values(1)
confere2.loc[:,'MC'] = confere2['VLRMRGBRT'] + confere2['VLRCSTMC']
confere2.loc['TOTAL',:] = confere2.sum()
print(confere2[['VLRVNDFATLIQ', 'VLRRCTLIQAPU', 'VLRMRGBRT', 'VLRCSTMC', 'MC']].to_string())

#CATEGORIA X FORNECEDOR
confere5 = df_full.pivot_table(index=['DESDRTCLLATU', 'DESCTGPRD', 'DESDIVFRN', 'MEDIDA'], values='DRIVER', aggfunc=sum).unstack()
confere5.columns = confere5.columns.get_level_values(1)
confere5.loc[:,'MC'] = confere5['VLRMRGBRT'] + confere5['VLRCSTMC']

#Gera dataset OMR_COMPRAS_Final
df_full = df_full.pivot_table(index=['NOMMES', 'CODGRPPRD', 'CODCTGPRD', 'CODSUBCTGPRD', 'CODDIVFRN', 'DESDIVFRN', 'CODESTUNI', 'DESTIPCNLVNDOMR', 'DESCTGPRD', 'DESDRTCLLATU'], columns=['MEDIDA'], values='DRIVER', aggfunc=sum).reset_index()
df_full = df_full.query('VLRVNDFATLIQ>0')
df_full.eval('VLRMRGCRB=VLRMRGBRT+VLRCSTMC', inplace=True)
df_full.reset_index(drop=True, inplace=True)
df_full.to_feather(caminho + 'bd/OMR_COMPRAS_Final.ft')

#Gera arquivo formato tabela de carga dwh.RLCOMRCMPOCDOPE
df_full = df_full.groupby(['NOMMES','CODGRPPRD','CODCTGPRD','CODSUBCTGPRD','CODDIVFRN','CODESTUNI','DESTIPCNLVNDOMR'])[['VLRVNDFATLIQ','VLRRCTLIQAPU','VLRMRGBRT','VLRMRGCRB']].sum().reset_index()
df_full.rename(columns={'CODDIVFRN':'CODFRN'}, inplace=True)

print('\n', "==> BASE CARGA POR TIPO DE CANAL")
confere3 = df_full.groupby(['DESTIPCNLVNDOMR'])[['VLRVNDFATLIQ','VLRRCTLIQAPU','VLRMRGBRT','VLRMRGCRB']].sum()
confere3.loc['TOTAL',:] = confere3.sum()
print(confere3.to_string())

print('\n', "TABELA DE CARGA")
print(df_full.nlargest(5, 'VLRVNDFATLIQ').to_string(index=False),'\n')
print(df_full.nsmallest(5, 'VLRVNDFATLIQ').to_string(index=False),'\n')
print(df_full.describe().to_string())

'''
==> CONFERENCIAS
Atualiza arquivo Confere_CARGARLC.xlsx
'''
#Confere TIPO CANAL PLANILHA DE CONTRIBUIÇÃO
confere4 = df_tipcnl.pivot_table(index=['DESTIPCNLVNDOMR','MEDIDA'], values='VALOR', aggfunc=sum).unstack()
confere4.columns = confere4.columns.get_level_values(1)
confere4.loc[:,'MC'] = confere4['VLRMRGBRT'] + confere4['VLRCSTMC']
confere4.loc['TOTAL',:] = confere4.sum()

#Estado Base Carga
confere6 = df_full.groupby(['CODESTUNI'])[['VLRVNDFATLIQ','VLRRCTLIQAPU','VLRMRGBRT','VLRMRGCRB']].sum()
confere6.loc['TOTAL',:] = confere6.sum()

#Atualiza arquivo excel
print('iniciando atualização do arquivo xlsx...')
from openpyxl import load_workbook
book = load_workbook(caminho + 'Confere_CARGARLC.xlsx')
writer = pd.ExcelWriter(caminho + 'Confere_CARGARLC.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

planilha = 'DIRETORIA'
ws = book[planilha]
for row in ws['A3:G19']:
  for cell in row:
    cell.value = None
dfxls = confere2[['VLRVNDFATLIQ', 'VLRRCTLIQAPU', 'VLRMRGBRT', 'VLRCSTMC', 'MC']].reset_index()
dfxls.to_excel(writer, sheet_name=planilha, startcol=0, startrow=2, header=False, index=False)
for row in ws['A22:G39']:
  for cell in row:
    cell.value = None
dfxls = confere[['VLRVNDFATLIQ', 'VLRRCTLIQAPU', 'VLRMRGBRT', 'VLRCSTMC', 'MC']].reset_index()
dfxls.to_excel(writer, sheet_name=planilha, startcol=0, startrow=21, header=False, index=False)

planilha = 'CANAL'
ws = book[planilha]
for row in ws['A3:G19']:
  for cell in row:
    cell.value = None
dfxls = confere3.reset_index()
dfxls.to_excel(writer, sheet_name=planilha, startcol=0, startrow=2, header=False, index=False)
for row in ws['A22:G39']:
  for cell in row:
    cell.value = None
dfxls = confere4[['VLRVNDFATLIQ', 'VLRRCTLIQAPU', 'VLRMRGBRT', 'MC']].reset_index()
dfxls.to_excel(writer, sheet_name=planilha, startcol=0, startrow=21, header=False, index=False)

planilha = 'FORNECEDOR'
ws = book[planilha]
for row in ws['A3:G2000']:
  for cell in row:
    cell.value = None
dfxls = confere5[['VLRVNDFATLIQ', 'VLRRCTLIQAPU', 'VLRMRGBRT', 'MC']].reset_index()
dfxls.to_excel(writer, sheet_name=planilha, startcol=0, startrow=2, header=False, index=False)

planilha = 'ESTADO'
ws = book[planilha]
for row in ws['A3:G30']:
  for cell in row:
    cell.value = None
dfxls = confere6.reset_index().sort_values(by='VLRVNDFATLIQ', ascending=False)
dfxls.to_excel(writer, sheet_name=planilha, startcol=0, startrow=2, header=False, index=False)

planilha = 'TABLE'
ws = book[planilha]
for row in ws['A2:N50']:
  for cell in row:
    cell.value = None
dfxls = df_full.nlargest(5, 'VLRVNDFATLIQ')
dfxls.to_excel(writer, sheet_name=planilha, startcol=0, startrow=1, header=True, index=False)
dfxls = df_full.nsmallest(5, 'VLRVNDFATLIQ')
dfxls.to_excel(writer, sheet_name=planilha, startcol=0, startrow=9, header=True, index=False)

writer.save()
book.close()

df_full.to_feather(caminho + 'bd/RLCOMRCMPOCDOPE_CARGA.ft')

print("Confere_CARGARLC.xlsx atualizado!")
print("Dataset gerado: OMR_COMPRAS_Final.ft")
print("Dataset gerado: RLCOMRCMPOCDOPE_CARGA.ft")